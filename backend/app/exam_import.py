"""
Reading a school's question spreadsheet.

The whole design rests on one observation about who is using this: a teacher
who uploads sixty questions and gets back "upload failed" gives up and goes
back to photocopying. A teacher who gets back "rows 14, 31 and 47 have
problems, here is what is wrong with each" fixes three cells and carries on.

So this parser **never fails on the first bad row**. It reads the entire sheet,
collects every problem with its row number, and reports them together. The
caller decides whether to import the good rows anyway.

Everything else here is about being forgiving of how spreadsheets actually
arrive from a busy teacher: stray whitespace, headers in different cases,
`Option A` instead of `option_a`, the answer written out in full instead of as
a letter, and forty empty rows at the bottom where someone once pressed delete.
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field

REQUIRED_COLUMNS = ("question", "option_a", "option_b", "option_c", "option_d", "correct")
OPTIONAL_COLUMNS = ("subject", "topic", "explanation", "image_url")

# Accepted header spellings. Teachers write "Option A", "OPTION_A", "optionA";
# all of them mean the same thing and none of them is a mistake worth
# rejecting a file over.
HEADER_ALIASES = {
    "question": {"question", "questions", "questiontext", "question_text", "q"},
    "option_a": {"option_a", "optiona", "a", "choicea", "choice_a"},
    "option_b": {"option_b", "optionb", "b", "choiceb", "choice_b"},
    "option_c": {"option_c", "optionc", "c", "choicec", "choice_c"},
    "option_d": {"option_d", "optiond", "d", "choiced", "choice_d"},
    "correct": {"correct", "answer", "correctoption", "correct_option", "correctanswer", "correct_answer", "ans"},
    "subject": {"subject", "subjects"},
    "topic": {"topic", "topics"},
    "explanation": {"explanation", "explanations", "solution", "reason"},
    "image_url": {"image_url", "image", "imageurl", "diagram"},
}

MAX_ROWS = 500


@dataclass
class RowError:
    row: int          # 1-based, matching what the teacher sees in Excel
    problem: str


@dataclass
class ImportResult:
    questions: list[dict] = field(default_factory=list)
    errors: list[RowError] = field(default_factory=list)
    # Header problems stop everything, because without headers no row can be read.
    fatal: str | None = None

    @property
    def ok(self) -> bool:
        return self.fatal is None and bool(self.questions)


def _normalise_header(value) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value or "").strip().lower())


def _map_headers(header_row) -> tuple[dict[str, int], str | None]:
    """Map canonical column name -> column index, tolerating spelling variants."""
    found: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        key = _normalise_header(cell)
        if not key:
            continue
        for canonical, aliases in HEADER_ALIASES.items():
            if key in {_normalise_header(a) for a in aliases} and canonical not in found:
                found[canonical] = idx
                break

    missing = [c for c in REQUIRED_COLUMNS if c not in found]
    if missing:
        return found, (
            "The spreadsheet is missing these columns: " + ", ".join(missing) +
            ". The first row must be the column headings."
        )
    return found, None


def _clean(value) -> str:
    """
    Excel gives back numbers, dates and None as well as strings.

    `1.0` for an option that the teacher typed as `1` is the common one, and
    showing a student "1.0" as an answer choice looks broken.
    """
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _parse_correct(raw: str, options: dict[str, str]) -> tuple[str | None, str | None]:
    """
    Work out which option is the answer.

    Accepts a letter, and also the answer written out in full -- which is the
    single most common thing a teacher does instead of writing "B", and is far
    too ordinary a mistake to reject a file over.
    """
    value = raw.strip()
    if not value:
        return None, "no correct answer given"

    letter = value.upper().strip(" .()")
    if letter in ("A", "B", "C", "D"):
        return letter, None

    # Written out in full, e.g. "5x" where option_a is "5x".
    for key, text in options.items():
        if text and text.strip().lower() == value.lower():
            return key.upper(), None

    return None, f"correct answer '{raw}' is not A, B, C or D, and does not match any option"


def parse_workbook(data: bytes, default_subject: str | None = None) -> ImportResult:
    """
    Read an .xlsx file into question dicts, collecting every problem.

    Returns good rows AND bad rows. The caller shows the errors and offers to
    import what worked -- which is almost always what a teacher wants, since
    fifty-seven usable questions out of sixty is a usable paper.
    """
    result = ImportResult()

    try:
        from openpyxl import load_workbook
    except ImportError:  # pragma: no cover - dependency is declared
        result.fatal = "Excel support is not installed on the server."
        return result

    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        sheet = wb.active
    except Exception:
        result.fatal = (
            "That file could not be opened as a spreadsheet. Save it as .xlsx "
            "from Excel or Google Sheets and try again."
        )
        return result

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        result.fatal = "The spreadsheet is empty."
        return result

    columns, header_error = _map_headers(rows[0])
    if header_error:
        result.fatal = header_error
        return result

    def cell(row, name: str) -> str:
        idx = columns.get(name)
        if idx is None or idx >= len(row):
            return ""
        return _clean(row[idx])

    seen_questions: set[str] = set()

    for offset, row in enumerate(rows[1:], start=2):   # start=2: row 1 is headers
        if len(result.questions) >= MAX_ROWS:
            result.errors.append(RowError(offset, f"stopped at the {MAX_ROWS}-question limit"))
            break

        # Trailing empty rows are normal in a spreadsheet, not an error worth
        # reporting -- a teacher who deleted content leaves hundreds of them.
        if not any(_clean(v) for v in row):
            continue

        text = cell(row, "question")
        options = {
            "a": cell(row, "option_a"),
            "b": cell(row, "option_b"),
            "c": cell(row, "option_c"),
            "d": cell(row, "option_d"),
        }

        if not text:
            result.errors.append(RowError(offset, "no question text"))
            continue

        blank = [k.upper() for k, v in options.items() if not v]
        if blank:
            result.errors.append(RowError(offset, f"option {', '.join(blank)} is empty"))
            continue

        correct, problem = _parse_correct(cell(row, "correct"), options)
        if problem:
            result.errors.append(RowError(offset, problem))
            continue

        fingerprint = re.sub(r"\s+", " ", text.lower())
        if fingerprint in seen_questions:
            result.errors.append(RowError(offset, "duplicate of an earlier question"))
            continue
        seen_questions.add(fingerprint)

        result.questions.append({
            "subject": cell(row, "subject") or default_subject,
            "topic": cell(row, "topic") or None,
            "question_text": text,
            "image_url": cell(row, "image_url") or None,
            "option_a": options["a"],
            "option_b": options["b"],
            "option_c": options["c"],
            "option_d": options["d"],
            "correct_option": correct,
            "explanation": cell(row, "explanation") or None,
        })

    if not result.questions and not result.fatal:
        result.fatal = (
            "No usable questions were found. Check that the first row holds the "
            "column headings and that there is at least one question below it."
        )

    return result


def template_rows() -> list[list[str]]:
    """
    The blank template a school downloads.

    Two worked examples rather than one, and the second is a negation question,
    because "which of the following is NOT" is where teachers most often put
    the answer in the wrong column.
    """
    return [
        list(REQUIRED_COLUMNS) + list(OPTIONAL_COLUMNS),
        [
            "Simplify 3x + 2x", "5x", "6x", "x", "5", "A",
            "Mathematics", "Algebra", "Add the coefficients: 3 + 2 = 5, so 3x + 2x = 5x.", "",
        ],
        [
            "Which of the following is NOT a prime number?", "2", "9", "7", "11", "B",
            "Mathematics", "Numbers", "9 = 3 x 3, so it has a factor other than 1 and itself.", "",
        ],
    ]
